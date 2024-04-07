from openpyxl import load_workbook
wb = load_workbook('cell_test.xlsx')
sheet = wb['書單1']
# 根據位置取得儲存格
c = sheet['A4']
# 取得儲存格資料
print(c.value)
# 修改資料
c.value = "App Inventor"
print(c.value)
# 以行號、列號指定儲存格
c = sheet.cell(row=4, column=4)
print(c.value)
# 取得指定範圍內儲存格物件
cellRange = sheet['A2':'A8']
# 以 for 迴圈逐一處理每個儲存格
for row in cellRange:
    for c in row:
        print(c.value)
#將修改過的活頁簿內容以另一個檔名儲存
wb.save('cell_test1.xlsx')
