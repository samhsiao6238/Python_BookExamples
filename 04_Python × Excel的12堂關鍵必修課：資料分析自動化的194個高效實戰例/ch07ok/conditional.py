from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

wb = load_workbook('score.xlsx')
target=wb.active

fail=CellIsRule(operator="lessThan",formula=[60],
                stopIfTrue=True,fill=PatternFill(
                "solid",start_color="FF0000",end_color="FF0000")
)
target.conditional_formatting.add("B2:B8",fail)
                
wb.save('score_ok.xlsx')
