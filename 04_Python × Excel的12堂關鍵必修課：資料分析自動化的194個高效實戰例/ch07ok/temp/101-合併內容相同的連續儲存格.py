from openpyxl import load_workbook
workbook = load_workbook('訂單金額表.xlsx')
worksheet = workbook['工作表1']
lists = []
num = 2
while True:
    datas = worksheet.cell(num, 1).value
    if datas:
        lists.append(datas)
    else:
        break
    num += 1
s = 0
e = 0
data = lists[0]
for m in range(len(lists)):
    if lists[m] != data:
        data = lists[m]
        e = m - 1
        if e >= s:
            worksheet.merge_cells(f'A{s + 2}:A{e + 2}')
            s = e + 1
    if m == len(lists) - 1:
        e = m
        worksheet.merge_cells(f'A{s + 2}:A{e + 2}')
workbook.save('訂單金額表1.xlsx')
