from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
workbook = load_workbook('訂單表5.xlsx')
worksheet = workbook['總表']
worksheet.merge_cells('A1:I1')
worksheet['A1'].font = Font(name='微軟正黑體', size=18, bold=True)
worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
worksheet.row_dimensions[1].height = 30
workbook.save('訂單表6.xlsx')
