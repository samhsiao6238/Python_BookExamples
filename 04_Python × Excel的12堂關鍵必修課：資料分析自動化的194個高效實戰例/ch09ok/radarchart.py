import openpyxl
from openpyxl.chart import RadarChart, Reference

wb=openpyxl.load_workbook("stock1.xlsx") #載入Excel活頁簿檔案
target=wb.active #將作用工作表內容設定給target變數
#設定要繪製圖表的資料參考範圍
price=Reference(target,min_col=3,max_col=6,min_row=1,max_row=target.max_row)
#設定要繪製圖表的分類參考範圍
stock_sort=Reference(target,min_col=2,max_col=2,min_row=2,max_row=target.max_row)
chart=RadarChart()  #建立圖
chart.grouping="stacked"
chart.title="股票獲利績效"  #統計圖表的標題名稱
#如果多設定chart.type="filled", 則會在每一個多邊形的內部區域填滿色彩
#將資料參考範圍加入圖表, 並令第一列為圖示名稱
chart.add_data(price,titles_from_data=True)
#新增類別物件，以作為圖表的分類
chart.set_categories(stock_sort)
#將圖表插入工作表中的指定儲存格位置
target.add_chart(chart,"A10")
#將程式的執行結果以另外一個新檔名加以儲存
wb.save("stock_radarchart.xlsx")
